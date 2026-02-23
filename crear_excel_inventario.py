# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inventario"

headers = [
    "Nº", "NOMBRE GENERICO", "NOMBRE COMER.", "PRESENTA.", "FOR. DE AUT. Nº",
    "FECHA", "EXISTENCIA ANTERIOR", "ENTRADA", "SALIDA", "JUSTIFICACION", "SALDO"
]
for col, h in enumerate(headers, 1):
    ws.cell(1, col, h)
    ws.cell(1, col).font = Font(bold=True)
    ws.cell(1, col).alignment = Alignment(horizontal="center", wrap_text=True)

data = [
    # Filas 1-15 (primera imagen)
    [1, "AMITRIPTILIN 25", "QUIMFA", "comp.", "1633230", "19/01/2026", 0, 119300, 0, "DISTRIB.PAIS", 119300],
    [2, "AMITRIPTILIN 25", "QUIMFA", "comp.", "1633247", "29/01/2026", 119300, 700, 114300, "DISTRIB.PAIS", 5700],
    [3, "ALPRAZOLAN 1", "", "comp", "", "", 410640, 0, 406640, "DISTRIB.PAIS", 4000],
    [4, "BUPROPION 150", "DALLAS", "comp", "1633215", "09/01/2026", 460, 5000, 2170, "DISTRIB.PAIS", 3290],
    [5, "CODEINA +PARA", "", "Comp", "", "", 633450, 0, 155150, "DISTRIB.PAIS", 478300],
    [6, "CLONAZEPAN 2,5", "CLOPAN", "Gotas", "1633239", "26/01/2026", 27111, 10004, 0, "DISTRIB.PAIS", 37115],
    [7, "CLONAZEPAN 2,5", "ETICOS", "Gotas", "1633228", "19/01/2026", 37115, 2022, 11537, "DISTRIB.PAIS", 27600],
    [8, "CLONAZEPAN 2", "", "comp", "", "", 879710, 0, 318880, "DISTRIB.PAIS", 560830],
    [9, "DIAZEPAN 10", "", "comp", "", "", 20690, 0, 5970, "DISTRIB.PAIS", 14720],
    [10, "DIAZEPAN 10", "", "Amp.", "", "", 1843, 0, 1520, "DISTRIB.PAIS", 323],
    [11, "FENOBARB. 100", "", "comp", "", "", 10, 0, 10, "DISTRIB.PAIS", 0],
    [12, "HALOP. DECA", "", "Amp.", "", "", 6, 0, 6, "DISTRIB.PAIS", 0],
    [13, "HALOPERIDOL 5", "QUIMFA", "Amp.", "1633233", "21/01/2026", 5681, 2022, 2535, "DISTRIB.PAIS", 5168],
    [14, "HALOPERIDOL 5", "", "COMP", "", "", 216750, 0, 48630, "DISTRIB.PAIS", 168120],
    [15, "HALOPERIDOL 2", "ANACTIVAN", "Gotas", "1633249", "30/01/2026", 221, 792, 221, "DISTRIB.PAIS", 792],
    # Filas 16-28 (segunda imagen)
    [16, "KETAMINA 50", "", "Amp.", "", "", 7608, 0, 3705, "DISTRIB.PAIS", 3903],
    [17, "LEVOMEPRO 25", "RELAXAM", "COMP", "1633220", "15/01/2026", 96070, 80000, 0, "DISTRIB.PAIS", 176070],
    [18, "LEVOMEPRO 25", "ETICOS", "COMP", "1633229", "19/01/2026", 176070, 83760, 0, "DISTRIB.PAIS", 259830],
    [19, "LEVOMEPRO 25", "ANSIOTEM", "COMP", "1633242", "27/01/2026", 259830, 120000, 131670, "DISTRIB.PAIS", 248160],
    [20, "MIDAZOLAN 5", "SEDANCOR", "Amp.", "1633241", "27/01/2026", 157776, 131660, 78705, "DISTRIB.PAIS", 210731],
    [21, "OLANZAPINA 10", "VTE SCAV.", "Comp", "1633235", "22/01/2026", 412650, 65500, 219970, "DISTRIB.PAIS", 258180],
    [22, "RISPERIDONA 3", "SICODONA", "COMP", "1633218", "14/01/2026", 740570, 4300, 0, "DISTRIB.PAIS", 744870],
    [23, "RISPERIDONA 3", "HEISECKE", "COMP", "1633232", "21/01/2026", 744870, 275000, 0, "DISTRIB.PAIS", 1019870],
    [24, "RISPERIDONA 3", "SICODONA", "COMP", "1633247", "29/01/2025", 1019870, 46270, 0, "DISTRIB.PAIS", 1066140],
    [25, "RISPERIDONA 3", "MIRABRIL", "COMP", "1633249", "30/01/2025", 1066140, 165000, 300980, "DISTRIB.PAIS", 930160],
    [26, "RISPERIDONA 1", "VIDACIL", "GTAS", "1633231", "19/01/2026", 3660, 3480, 0, "DISTRIB.PAIS", 7140],
    [27, "RISPERIDONA 1", "GUAYAKI", "GTAS", "1633221", "21/01/2025", 7140, 406, 665, "DISTRIB.PAIS", 6881],
    [28, "RISPERIDONA 1", "SICODONA", "GTAS", "1633247", "29/01/2026", 6881, 3174, 3174, "DISTRIB.PAIS", 6881],
    # Filas 29-47 (tercera imagen)
    [29, "ALFENTANILO", "RAPIFEN", "Amp.", "1633243", "28/01/2026", 5620, 4500, 0, "DISTRIB.PAIS", 10120],
    [30, "ALFENTANILO", "VTE SCAVO", "Amp.", "1633235", "22/01/2026", 10120, 8000, 4453, "DISTRIB.PAIS", 13667],
    [31, "FENTANILO 0,05", "FENTANILO", "Amp.", "1633210", "02/01/2026", 4600, 5000, 0, "DISTRIB.PAIS", 9600],
    [32, "FENTANILO 0,05", "FENTANILO", "Amp.", "1633212", "05/01/2026", 9600, 5000, 0, "DISTRIB.PAIS", 14600],
    [33, "FENTANILO 0,05", "INTERLABO", "Amp.", "1633211", "05/01/2026", 14600, 5200, 0, "DISTRIB.PAIS", 19800],
    [34, "FENTANILO 0,05", "INTERLABO", "Amp.", "1633213", "08/01/2026", 19800, 6200, 0, "DISTRIB.PAIS", 26000],
    [35, "FENTANILO 0,05", "FENTANILO", "Amp.", "1633214", "08/01/2026", 26000, 5000, 0, "DISTRIB.PAIS", 31000],
    [36, "FENTANILO 0,05", "FENTANILO", "Amp.", "1633216", "13/01/2026", 31000, 5000, 0, "DISTRIB.PAIS", 36000],
    [37, "FENTANILO 0,05", "BIOSANO", "Amp.", "1633219", "14/01/2026", 36000, 7200, 0, "DISTRIB.PAIS", 43200],
    [38, "FENTANILO 0,05", "FENTANILO", "Amp.", "1633234", "21/01/2026", 43200, 5000, 0, "DISTRIB.PAIS", 48200],
    [39, "FENTANILO 0,05", "INTERLABO", "Amp.", "1633237", "23/01/2026", 48200, 5500, 0, "DISTRIB.PAIS", 53700],
    [40, "FENTANILO 0,05", "VOLDEX", "Amp.", "1633240", "28/01/2026", 48700, 7517, 0, "DISTRIB.PAIS", 56217],
    [41, "FENTANILO 0,05", "FENTANILO", "Amp.", "1633244", "28/01/2026", 56217, 5000, 0, "DISTRIB.PAIS", 61217],
    [42, "FENTANILO 0,05", "PROSALUD", "Amp.", "1633246", "28/01/2026", 61217, 67475, 0, "DISTRIB.PAIS", 128692],
    [43, "FENTANILO 0,05", "BIOSANO", "Amp.", "1633248", "29/01/2026", 128692, 6100, 6000, "DISTRIB.PAIS", 128792],
    [44, "METILFENIDATO", "ETICOS", "COMP", "1633238", "23/01/2026", 36070, 34550, 11500, "DISTRIB.PAIS", 59120],
    [45, "MORF. CLORH 30", "C.X 50", "", "", "", 563, 0, 221, "DISTRIB.PAIS", 342],
    [46, "MORF. CLORH 10", "FUSA", "Fco am", "1633217", "14/01/2026", 28161, 20504, 16599, "DISTRIB.PAIS", 32066],
    [47, "TALIDOMIDA 100", "PROMEPAR", "Comp.", "1633226", "19/01/2026", 23440, 6560, 0, "DISTRIB.PAIS", 30000],
]

for row_idx, row in enumerate(data, 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row_idx, col_idx, value)

# Ajustar ancho de columnas
for col in ws.columns:
    max_len = max(len(str(c.value or "")) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 1, 18)

wb.save("Inventario_farmacos.xlsx")
print("Archivo Inventario_farmacos.xlsx creado correctamente.")
